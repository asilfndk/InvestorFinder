"""
Export routes for downloading investor data as CSV or Excel.
"""

import csv
import io
from typing import List
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
import logging

from app.models import InvestorProfile, InvestorSearchRequest
from app.services import ChatService, InvestorService
from app.services.memory_service import get_memory_service

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/export", tags=["export"])

# Service dependencies
_chat_service: ChatService = None
_investor_service: InvestorService = None


async def get_chat_service() -> ChatService:
    """Dependency for getting chat service."""
    global _chat_service
    if _chat_service is None:
        _chat_service = ChatService()
        await _chat_service.initialize()
    return _chat_service


async def get_investor_service() -> InvestorService:
    """Dependency for getting investor service."""
    global _investor_service
    if _investor_service is None:
        _investor_service = InvestorService()
    return _investor_service


def investors_to_csv(investors: List[InvestorProfile]) -> str:
    """Convert investors list to CSV string."""
    output = io.StringIO()

    fieldnames = [
        "name", "title", "company", "email", "linkedin_url",
        "location", "bio", "investment_focus", "source"
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for inv in investors:
        row = {
            "name": inv.name or "",
            "title": inv.title or "",
            "company": inv.company or "",
            "email": inv.email or "",
            "linkedin_url": inv.linkedin_url or "",
            "location": inv.location or "",
            "bio": (inv.bio[:200] + "...") if inv.bio and len(inv.bio) > 200 else (inv.bio or ""),
            "investment_focus": ", ".join(inv.investment_focus) if inv.investment_focus else "",
            "source": inv.source or ""
        }
        writer.writerow(row)

    return output.getvalue()


def investors_to_excel_bytes(investors: List[InvestorProfile]) -> bytes:
    """Convert investors list to Excel bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Investors"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5",
                              end_color="4F46E5", fill_type="solid")

    # Headers
    headers = [
        "Name", "Title", "Company", "Email", "LinkedIn URL",
        "Location", "Bio", "Investment Focus", "Source"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, inv in enumerate(investors, 2):
        ws.cell(row=row_idx, column=1, value=inv.name or "")
        ws.cell(row=row_idx, column=2, value=inv.title or "")
        ws.cell(row=row_idx, column=3, value=inv.company or "")
        ws.cell(row=row_idx, column=4, value=inv.email or "")
        ws.cell(row=row_idx, column=5, value=inv.linkedin_url or "")
        ws.cell(row=row_idx, column=6, value=inv.location or "")
        ws.cell(row=row_idx, column=7, value=(
            inv.bio[:500] + "...") if inv.bio and len(inv.bio) > 500 else (inv.bio or ""))
        ws.cell(row=row_idx, column=8, value=", ".join(
            inv.investment_focus) if inv.investment_focus else "")
        ws.cell(row=row_idx, column=9, value=inv.source or "")

    # Adjust column widths
    column_widths = [25, 30, 30, 35, 50, 25, 60, 40, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


@router.get("/{conversation_id}/csv")
async def export_conversation_csv(
    conversation_id: str,
    chat_service: ChatService = Depends(get_chat_service)
):
    """
    Export investors from a conversation as CSV.

    Returns a downloadable CSV file with all investors found in the conversation.
    """
    investors = chat_service.get_conversation_investors(conversation_id)

    if not investors:
        raise HTTPException(
            status_code=404, detail="No investors found for this conversation")

    csv_content = investors_to_csv(investors)

    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=investors_{conversation_id[:8]}.csv"
        }
    )


@router.get("/{conversation_id}/excel")
async def export_conversation_excel(
    conversation_id: str,
    chat_service: ChatService = Depends(get_chat_service)
):
    """
    Export investors from a conversation as Excel.

    Returns a downloadable Excel file with all investors found in the conversation.
    """
    investors = chat_service.get_conversation_investors(conversation_id)

    if not investors:
        raise HTTPException(
            status_code=404, detail="No investors found for this conversation")

    excel_bytes = investors_to_excel_bytes(investors)

    return StreamingResponse(
        iter([excel_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=investors_{conversation_id[:8]}.xlsx"
        }
    )


@router.post("/investors/csv")
async def export_investors_csv_direct(investors: List[InvestorProfile]):
    """
    Export a list of investors directly as CSV.

    Accepts a list of investor profiles and returns a CSV file.
    """
    if not investors:
        raise HTTPException(status_code=400, detail="No investors provided")

    csv_content = investors_to_csv(investors)

    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=investors.csv"
        }
    )


@router.post("/investors/excel")
async def export_investors_excel_direct(investors: List[InvestorProfile]):
    """
    Export a list of investors directly as Excel.

    Accepts a list of investor profiles and returns an Excel file.
    """
    if not investors:
        raise HTTPException(status_code=400, detail="No investors provided")

    excel_bytes = investors_to_excel_bytes(investors)

    return StreamingResponse(
        iter([excel_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=investors.xlsx"
        }
    )
